from subprocess import check_output
from locale import setlocale,LC_ALL
from tkinter import Tk,filedialog,messagebox,Button,font,Label,Canvas,ttk,StringVar
from functools import partial
from os.path import isdir,exists
from os import listdir,getcwd,remove,mkdir,chdir
from shutil import copyfile,rmtree
from ajl_tools import to_tuple,to_list
from subprocess import Popen
from regex import sub
from zipfile import ZipFile,ZIP_DEFLATED
import gc
from openpyxl import load_workbook

error_message = messagebox.showerror
yesno_message = messagebox.askquestion
confirm_message = messagebox.askokcancel
ttk_Checkbox = ttk.Checkbutton

setlocale(LC_ALL,'')

class Dynamic_Controller:

    def __init__(self):
        self.explicit_rerun = None
        self.sde_info = [None,None,None,None]
        self.datasources_created = False
        self.datasources_generated = False
        self.box_vals = [None for n in range(16)]
        self.backup_gdb = None
        self.update_master_dasid = None

    def append_box_vals(self,item :str,num : int):

        if item == 'NIHIL':
            self.box_vals[num] = None
        else:
            self.box_vals[num] = item

dc_object = Dynamic_Controller()

print("Initializing ArcPy module...")
try:
    from arcpy import env,da,management,SetLogHistory,SetLogMetadata,ListDatasets,ListFeatureClasses,Describe,ListFields,ListTables,Exists,conversion,Describe
except Exception:
    print("\nERROR!\nYou are not signed into an Esri account on this computer.\nAttempting to open ArcGIS Pro...")
    try:
        p = Popen("C:/Program Files/ArcGIS/Pro/bin/ArcGISPro.exe")
        print("Please close ArcGIS Pro after signing into your Esri account in order for the program to continue.")
        p.wait() ; p.terminate()
        from arcpy import env,da,management,SetLogHistory,SetLogMetadata,ListDatasets,ListFeatureClasses,Describe,ListFields,ListTables,Exists,conversion,Describe
    except:
        print("Unable to open ArcGIS Pro for you. Please open ArcGIS Pro and sign in before running this program again.")
        input("Press Enter/Return to close this program: ")
        exit()

mgnt_mfl = management.MakeFeatureLayer
mgnt_AddField = management.AddField
mgnt_lb_attribute = management.SelectLayerByAttribute
mgnt_lb_location = management.SelectLayerByLocation
mgnt_Delete = management.Delete
mgnt_DeleteRows = management.DeleteRows
mgnt_DeleteField = management.DeleteField
mgnt_DET = management.DisableEditorTracking
mgnt_EET = management.EnableEditorTracking
mgnt_CreateTable = management.CreateTable
mgnt_CreateFC = management.CreateFeatureclass

SetLogHistory(False) ; SetLogMetadata(False)

try:
    check_output('nvidia-smi')
    env.processorType = "GPU"
except Exception:
    env.processorType = "CPU"
    env.parallelProcessingFactor = "50%"

env.overwriteOutput = True

class Initialization_Variables:

    def __init__(self):

        self.das_table_gdb_path = None

class GDB_Info:

    def __init__(self,gdb_path):

        self.pnt_fc_names = []
        self.pln_fc_names = []
        self.poly_fc_names = []
        self.datasets = []
        self.mapunit_vals = []
        self.x_id_count = 0
        self.spatial_reference = []
        self.idRootDict = {"CartographicLines": "CAL","ContactsAndFaults":"CAF","CMULines":"CMULIN","CMUMapUnitPolys":"CMUMUP","CMUPoints":"CMUPNT","CMUText":"CMUTXT","DataSources":"DAS","DataSourcePolys":"DSP","DescriptionOfMapUnits":"DMU","ExtendedAttributes":"EXA","FossilPoints":"FSP","GenericPoints":"GNP","GenericSamples":"GNS","GeochemPoints":"GCM","GeochronPoints":"GCR","GeologicEvents":"GEE","GeologicLines":"GEL","Glossary":"GLO","IsoValueLines":"IVL","MapUnitPoints":"MPT","MapUnitPolys":"MUP","MapUnitOverlayPolys":"MUO","MiscellaneousMapInformation":"MMI","OrientationPoints":"ORP","OtherLines":"OTL","OverlayPolys":"OVP","PhotoPoints":"PHP","RepurposedSymbols":"RPS","Stations":"STA","StandardLithology":"STL","MapUnitPointAnno24k":"ANO"}
        self.tables = tuple(ListTables())

        for item in tuple(ListDatasets()):
            if "CrossSection" in item:
                self.datasets.append(item)
            elif "GeologicMap" == item:
                self.datasets.append(item)

        if "GeologicMap" in self.datasets:
            if len(self.datasets) > 1:
                self.datasets.remove("GeologicMap")
                self.datasets = sorted(self.datasets)
                self.datasets.insert(0,"GeologicMap")
            self.datasets = tuple(self.datasets)
            self.pnt_fc_names.append([]) ; self.pln_fc_names.append([]) ; self.poly_fc_names.append([])
            for fc in ListFeatureClasses(feature_dataset="GeologicMap"):
                if (shape_type := Describe(f"{gdb_path}/GeologicMap/{fc}").shapeType) == 'Point':
                    self.pnt_fc_names[-1].append(fc)
                elif shape_type in ('Line','Polyline'):
                    self.pln_fc_names[-1].append(fc)
                elif shape_type == 'Polygon':
                    self.poly_fc_names[-1].append(fc)
            self.pnt_fc_names[-1] = tuple(self.pnt_fc_names[-1]) ; self.pln_fc_names[-1] = tuple(self.pln_fc_names[-1]) ; self.poly_fc_names[-1] = tuple(self.poly_fc_names[-1])
            if len(self.datasets) > 1:
                for n in range(1,len(self.datasets)):
                    self.pnt_fc_names.append([])
                    self.pln_fc_names.append([])
                    self.poly_fc_names.append([])
                    for fc in ListFeatureClasses(feature_dataset=self.datasets[n]):
                        if (shape_type := Describe(f"{gdb_path}/{self.datasets[n]}/{fc}").shapeType) == 'Point':
                            self.pnt_fc_names[-1].append(fc)
                        elif shape_type in ('Line','Polyline'):
                            self.pln_fc_names[-1].append(fc)
                        elif shape_type == 'Polygon':
                            self.poly_fc_names[-1].append(fc)
            self.pnt_fc_names = to_tuple(self.pnt_fc_names) ; self.pln_fc_names = to_tuple(self.pln_fc_names) ; self.poly_fc_names = to_tuple(self.poly_fc_names)
        else:
            self.pnt_fc_names = None ; self.pln_fc_names = None ; self.poly_fc_names = None

        self.getSpatialReferences(gdb_path)

    def getSpatialReferences(self,gdb_path : str):

        for a in range(len(self.datasets)):
            if len(self.pnt_fc_names[a]):
                self.spatial_reference.append(da.Describe(f"{gdb_path}/{self.datasets[a]}/{self.pnt_fc_names[a][0]}")["spatialReference"].name)
            elif len(self.pln_fc_names[a]):
                self.spatial_reference.append(da.Describe(f"{gdb_path}/{self.datasets[a]}/{self.pln_fc_names[a][0]}")["spatialReference"].name)
            elif len(self.poly_fc_names[a]):
                self.spatial_reference.append(da.Describe(f"{gdb_path}/{self.datasets[a]}/{self.poly_fc_names[a][0]}")["spatialReference"].name)
            else:
                self.spatial_reference.append(None)

        self.spatial_reference = tuple(self.spatial_reference)

print("Initialization successful.\n")

root2 = Tk()

ws = root2.winfo_screenwidth() ; hs = root2.winfo_screenheight()

root2.update() ; root2.destroy()

gc.collect()

def main() -> None:

    valid_selection = False

    root3 = Tk() ; root3.withdraw()

    while not valid_selection:
        check_dir = filedialog.askdirectory(title="Select .gdb (Geodatabase) directory/folder")
        if isdir((check_dir := check_dir.replace("\\","/"))):
            if check_dir.endswith(".gdb"):
                dir_name = check_dir[:]
                while dir_name.find("/") != -1:
                    dir_name = dir_name[dir_name.find("/")+1:]
                print(f"{dir_name} has been selected.")
                break

    root3.update() ; root3.destroy()
    gc.collect()

    env.workspace = check_dir[:]

    gdb_info = GDB_Info(check_dir)

    def getRootName(fc_name : str) -> str:

        if fc_name.startswith("CS"):
            if fc_name[3:] in gdb_info.idRootDict.keys():
                return f"{fc_name[:3]}{gdb_info.idRootDict[fc_name[3:]]}"
            else:
                gdb_info.x_id_count += 1
                return f"{fc_name[:3]}X{gdb_info.x_id_count}X"
        else:
            if fc_name in gdb_info.idRootDict.keys():
                return gdb_info.idRootDict[fc_name]
            else:
                gdb_info.x_id_count += 1
                return f"X{gdb_info.x_id_count}X"

    if not "GeologicMap" in gdb_info.datasets:
        print("Non-GeMS geodatabase was selected!\nPlease choose a GeMS geodatabase.")
        main()

    init_vars = Initialization_Variables()

    yes_strings = {'true','t','yes','y','sure','okay','ok','k','yeah','alright','all right','correct','affirmative','right','valid','allow','permit','allowed','validate','validated','permitted','permited','confirm','confirmed','on','1','approved','approve','accept','accepted'}

    if exists("init.txt"):
        try:
            line_param = {line[:line.find("|")].lower() : line[line.find("|")+1:].rstrip("\n") for line in open("init.txt").readlines()}
            for item in line_param.keys():
                if "\\" in line_param[item]:
                    line_param[item] = line_param[item].replace("\\","/")
            if 'sde_server_type' in line_param.keys():
                dc_object.sde_info[0] = line_param['sde_server_type']
            if 'sde_instance' in line_param.keys():
                dc_object.sde_info[1] = line_param['sde_instance']
            if 'sde_authentication' in line_param.keys():
                dc_object.sde_info[2] = line_param['sde_authentication']
            if 'sde_database' in line_param.keys():
                dc_object.sde_info[3] = line_param['sde_database']
            del line_param
        except Exception:
            pass

    output_trans = {True: "ENABLED",False : "DISABLED"}

    del yes_strings ; del output_trans
    gc.collect()

    geometry_str = "800x800+%d+%d" % ((ws/2)-400,(hs/2)-400)

    root = Tk()
    root.title("Extra GeMS Tools")
    root.geometry(geometry_str)
    root.resizable(False,False)
    root.iconbitmap(f"{getcwd()}/_assets/ve.ico")
    helv_label = font.Font(family='Helvetica',size=17,weight=font.BOLD)
    helv_button = font.Font(family='Helvetica',size=11,slant=font.ITALIC)

    def editorErrorMessage():
        error_message("Unable to Edit","For unknown reasons, certain one or more\nfeature classes and/or tables are not allowing for\nalterations. This may be due to a \"rouge\" .LOCK file\nand/or GIS software accessing the same geodatabase while this\nprocess is running. Please close said software before\nrerunning this program!\n\nAn unaccounted error may also be resonsible for seeing this message.\n\nClick OK to close this program.")
        exit()

    def switchGDB():
        print("Changing selected geodatabase....")
        gc.collect()
        root.destroy()
        dc_object.backup_gdb = None
        main()

    def point_mapunit():

        env.workspace = check_dir[:]

        if dc_object.explicit_rerun:
            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
        else:
            print("Determining assigning geologic map unit values to MapUnit fields with point feature classes with said field...")

        try:
            for a in range(len(gdb_info.datasets)):
                prefix = "" ; error_str = None
                if gdb_info.datasets[a].startswith("CrossSection"):
                    prefix = f"CS{gdb_info.datasets[a][-1]}"
                if not "%sMapUnitPolys" % prefix in gdb_info.poly_fc_names[a]:
                    error_str = "Unable to find '%sMapUnitPolys' in '%s'" % (prefix,gdb_info.datasets[a])
                    if not len(gdb_info.pnt_fc_names[a]):
                        error_str = "%s and cannot find any point feature classes in '%s'." % (error_str,gdb_info.datasets[a])
                    else:
                        error_str = "%s." % error_str
                    print(error_str)
                elif not len(gdb_info.pnt_fc_names[a]):
                    print("Unable to find any point feature classes in '%s'." % gdb_info.datasets[a])
                else:
                    poly_lyr = mgnt_mfl(f"{gdb_info.datasets[a]}/{prefix}MapUnitPolys","mapunit_poly_lyr")
                    try:
                        mapunit_vals = tuple(sorted({row[0] for row in da.SearchCursor(f"{gdb_info.datasets[a]}/{prefix}MapUnitPolys",["MapUnit"]) if not row[0] is None}))
                    except:
                        error_message("Missing Field",f"'{prefix}MapUnitPolys' is missing 'MapUnit' field.")
                        if dc_object.explicit_rerun:
                            try:
                                edit.stopOperation()
                            except Exception:
                                pass
                            try:
                                edit.stopEditing(save_changes=True)
                            except Exception:
                                pass
                        mgnt_Delete(poly_lyr)
                        if dc_object.explicit_rerun:
                            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
                        continue
                    if not len(mapunit_vals):
                        if dc_object.explicit_rerun:
                            try:
                                edit.stopOperation()
                            except Exception:
                                pass
                            try:
                                edit.stopEditing(save_changes=True)
                            except Exception:
                                pass
                        mgnt_Delete(poly_lyr)
                        if dc_object.explicit_rerun:
                            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
                        continue
                    if not len((relevant_fc := tuple([gdb_info.pnt_fc_names[a][b] for b in range(len(gdb_info.pnt_fc_names[a])) if "MapUnit" in tuple([field.name for field in ListFields(f"{gdb_info.datasets[a]}/{gdb_info.pnt_fc_names[a][b]}")])]))):
                        if dc_object.explicit_rerun:
                            try:
                                edit.stopOperation()
                            except Exception:
                                pass
                            try:
                                edit.stopEditing(save_changes=True)
                            except Exception:
                                pass
                        mgnt_Delete(poly_lyr)
                        if dc_object.explicit_rerun:
                            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
                        continue
                    pnt_lyrs = []
                    for b in range(len(relevant_fc)):
                        mgnt_mfl(f"{gdb_info.datasets[a]}/{relevant_fc[b]}",f"pnt_lyr_{b}")
                        pnt_lyrs.append(f"pnt_lyr_{b}")
                    pnt_lyrs = tuple(pnt_lyrs)
                    oid_field_name = []
                    for item in relevant_fc:
                        for field in ListFields(f"{gdb_info.datasets[a]}/{item}"):
                            oid_field_name.append(field.name)
                            break
                    oid_field_name = tuple(oid_field_name)
                    pnt_objectid = [[] for _ in relevant_fc] ; pnt_mapunit = [[] for _ in relevant_fc]
                    for unit in mapunit_vals:
                        selected_polys,count = mgnt_lb_attribute(poly_lyr,"NEW_SELECTION",f"MapUnit = '{unit}'","NON_INVERT")
                        del count
                        for b in range(len(pnt_lyrs)):
                            selected_pnts,redundant,count = mgnt_lb_location(pnt_lyrs[b],"INTERSECT",selected_polys,"","NEW_SELECTION","NOT_INVERT")
                            del redundant
                            if count:
                                for row in da.SearchCursor(selected_pnts,[oid_field_name[b]]):
                                    pnt_objectid[b].append(row[0])
                                    pnt_mapunit[b].append(unit)
                            else:
                                continue
                    pnt_objectid = tuple(pnt_objectid) ; pnt_mapunit = tuple(pnt_mapunit)
                    for b in range(len(relevant_fc)):
                        with da.UpdateCursor(f"{gdb_info.datasets[a]}/{relevant_fc[b]}",[oid_field_name[b],"MapUnit"]) as cursor:
                            for row in cursor:
                                if row[0] in pnt_objectid[b]:
                                    row[1] = pnt_mapunit[b][pnt_objectid[b].index(row[0])]
                                cursor.updateRow(row)

                    if dc_object.explicit_rerun:
                        try:
                            edit.stopOperation()
                        except Exception:
                            pass
                        try:
                            edit.stopEditing(save_changes=True)
                        except Exception:
                            pass

                    if Exists(poly_lyr):
                        mgnt_Delete(poly_lyr)
                    for item in pnt_lyrs:
                        if Exists(item):
                            mgnt_Delete(item)
        except Exception:
            if not dc_object.explicit_rerun:
                print("Denied editing permissions for one or more feature classes.\nRerunning with explicit editing requests.\n")
                dc_object.explicit_rerun = True ; gc.collect() ; point_mapunit()
            else:
                editorErrorMessage()

        dc_object.explicit_rerun = False

        gc.collect()

    def resetFcId():

        env.workspace = check_dir[:]

        if dc_object.explicit_rerun:
            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
        else:
            print("Renumbering _ID fields in relevant feature classes...")

        def renum_ID_Fields(dataset : str,fc : str):
            id_field = None
            for fld in [field.name for field in ListFields(f"{dataset}/{fc}")]:
                if fld.endswith("_ID"):
                    id_field = fld[:]
                    break
            root_name = getRootName(fc)
            num_rows = 0 ; counter = 1
            for row in da.SearchCursor(f"{dataset}/{fc}",id_field):
                num_rows += 1
            num_rows_str_len = len(str(num_rows))
            with da.UpdateCursor(f"{dataset}/{fc}",id_field) as cursor:
                for row in cursor:
                    row[0] = "%s%s%i" % (root_name,"0" * (num_rows_str_len-len(str(counter))),counter)
                    cursor.updateRow(row)
                    counter += 1

        try:
            for a in range(len(gdb_info.datasets)):
                for b in range(len(gdb_info.pnt_fc_names[a])):
                    renum_ID_Fields(gdb_info.datasets[a],gdb_info.pnt_fc_names[a][b])
                for b in range(len(gdb_info.pln_fc_names[a])):
                    renum_ID_Fields(gdb_info.datasets[a],gdb_info.pln_fc_names[a][b])
                for b in range(len(gdb_info.poly_fc_names[a])):
                    renum_ID_Fields(gdb_info.datasets[a],gdb_info.poly_fc_names[a][b])

            for gems_table in ("DescriptionOfMapUnits","Glossary"):
                if gems_table in gdb_info.tables:
                    id_field = None
                    for fld in [field.name for field in ListFields(gems_table)]:
                        if fld.endswith("_ID"):
                            id_field = fld[:]
                            break
                    root_name = getRootName(gems_table)
                    num_rows = 0 ; counter = 1
                    for row in da.SearchCursor(gems_table,[id_field]):
                        num_rows += 1
                    num_rows_str_len = len(str(num_rows))
                    with da.UpdateCursor(gems_table,[id_field]) as cursor:
                        for row in cursor:
                            row[0] = "%s%s%i" % (root_name,"0" * (num_rows_str_len-len(str(counter))),counter)
                            cursor.updateRow(row)
                            counter += 1

            if dc_object.explicit_rerun:
                try:
                    edit.stopOperation()
                except Exception:
                    pass
                try:
                    edit.stopEditing(save_changes=True)
                except Exception:
                    pass

            print("Process completed successfully.\n")
        except Exception:
            if not dc_object.explicit_rerun:
                print("Denied editing permissions for one or more feature classes and/or tables.\nRerunning with explicit editing requests.\n")
                dc_object.explicit_rerun = True ; gc.collect() ; resetFcId()
            else:
                editorErrorMessage()

        dc_object.explicit_rerun = False

        gc.collect()

    def removeETfields():

        env.workspace = check_dir[:]

        if not confirm_message("Confirm Editor Tracking Removal","These changes cannot be reverted and Editor Tracking related data be restored. Proceed?"):
            print("\nRemoval Editor Tracking related data aborted!\n")
            return

        if dc_object.explicit_rerun:
            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
        else:
            print("Disabling editor tracking for all feature classes and tables...")

        ET_fields = {'created_user','created_date','last_edited_user','last_edited_date'}

        def delete_ET_Fields(dataset : str, fc : str):
            try:
                mgnt_DET(f"{dataset}/{fc}")
            except Exception:
                pass
            if len((del_fields := tuple([field.name for field in ListFields(f"{dataset}/{fc}") if field.name in ET_fields]))):
                mgnt_DeleteField(f"{dataset}/{fc}",del_fields)

        try:
            for a in range(len(gdb_info.datasets)):
                for b in range(len(gdb_info.pnt_fc_names[a])):
                    delete_ET_Fields(gdb_info.datasets[a],gdb_info.pnt_fc_names[a][b])
                for b in range(len(gdb_info.pln_fc_names[a])):
                    delete_ET_Fields(gdb_info.datasets[a],gdb_info.pln_fc_names[a][b])
                for b in range(len(gdb_info.poly_fc_names[a])):
                    delete_ET_Fields(gdb_info.datasets[a],gdb_info.poly_fc_names[a][b])

            if len(gdb_info.tables):
                for a in range(len(gdb_info.tables)):
                    mgnt_DET(gdb_info.tables[a])
                    if len((del_fields := tuple([field.name for field in ListFields(gdb_info.tables[a]) if field.name in ET_fields]))):
                        mgnt_DeleteField(gdb_info.tables[a],del_fields)

            if dc_object.explicit_rerun:
                try:
                    edit.stopOperation()
                except Exception:
                    pass
                try:
                    edit.stopEditing(save_changes=True)
                except Exception:
                    pass
        except Exception:
            if not dc_object.explicit_rerun:
                print("Denied editing permissions for one or more feature classes and/or tables.\nRerunning with explicit editing requests...\n")
                dc_object.explicit_rerun = True ; gc.collect() ; removeETfields()
            else:
                editorErrorMessage()

        dc_object.explicit_rerun = False

        gc.collect()

    def end_window():
        print("Closing program....")
        root.destroy()

    def alphabetizeGlossary():

        env.workspace = check_dir[:]

        if dc_object.explicit_rerun:
            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
        else:
            print("\nOrganizing Glossary terms in alphabetical order...")

        if not "Glossary" in gdb_info.tables:
            error_message("Missing Glossary","No Glossary table was found!")
        else:
            required_fields = ('Term','Definition','DefinitionSourceID')
            if not dc_object.explicit_rerun:
                counter = 0
                for item in tuple([field.name for field in ListFields('Glossary',field_type='String')]):
                    if item in required_fields:
                        counter += 1
                if counter != 3:
                    error_message("Missing Fields","One or more of the following fields in Glossary is/are: Term, Definition,DefinitionSourceID")
                    return
                del counter
            try:
                glossary_term = [] ; glossary_def = [] ; unknown_counter = 1
                for row in da.SearchCursor("Glossary",required_fields):
                    if row[0] is None:
                        glossary_term.append(f"zzz_UNKNOWN_{unknown_counter}")
                        unknown_counter += 1
                    elif row[0].replace(' ','') == '':
                        glossary_term.append(f'zzz_UNKNOWN_{unknown_counter}')
                        unknown_counter += 1
                    else:
                        glossary_term.append(row[0])
                    glossary_def.append((row[1],row[2]))
                del unknown_counter
                glossary_term = tuple(glossary_term) ; glossary_def = tuple(glossary_def) ; ordered = tuple(sorted(glossary_term,key=str.lower)) ; counter = 0
                with da.UpdateCursor("Glossary",required_fields) as cursor:
                    for row in cursor:
                        row[0] = glossary_term[(num_index := glossary_term.index(ordered[counter]))]
                        row[1] = glossary_def[num_index][0]
                        row[2] = glossary_def[num_index][1]
                        counter += 1
                        cursor.updateRow(row)

            except Exception:
                if not dc_object.explicit_rerun:
                    print("Denied editing permissions for Glossary table.\nRerunning with explicit editing requests...\n")
                    dc_object.explicit_rerun = True ; gc.collect() ; alphabetizeGlossary()
                else:
                    editorErrorMessage()

        dc_object.explicit_rerun = False

        gc.collect()

        print("Alphabetizing successful!\n")

    def autopopulateDAS():

        env.workspace = check_dir[:]

        relevant_fields = {'DataSourceID',"DataSources_ID",'DefinitionSourceID','LocationSourceID','OrientationSourceID'}

        if dc_object.update_master_dasid == None:
            user_response = yesno_message("Update via SDE?","Would you like to use the most up-to-date master list of DASID information in order to fill out your DataSources table? Select YES for this to work. Please note that this may take several minutes to complete. In addition, you will need to a valid ESRI/ArcGIS Pro account and internet connection for this to work. If you select NO, a pre-existing Microsoft Excel spreadsheet of said master list will be used instead.\n\nAfter selecting either YES or NO, this message will not appear again for the duration that this program is running.")
            if user_response == 'yes':
                dc_object.update_master_dasid = True
            else:
                dc_object.update_master_dasid = False
            del user_response

        if dc_object.explicit_rerun:
            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
        else:
            if dc_object.update_master_dasid:
                print("Attempting to find master list of DASID information (this may take some time depending upon internet connection speed as well as dependant upon factors related to the server)...")
        if dc_object.update_master_dasid:
            isConnected = True
            if exists('DataSources.xlsx'):
                if exists('_temp'):
                    rmtree('_temp')
                mkdir('_temp')
                copyfile('DataSources.xlsx','_temp/DataSources.xlsx')
                remove('DataSources.xlsx')
    
            if Exists(f'{getcwd()}/_assets/temp.gdb/DataSources'):
                mgnt_Delete(f'{getcwd()}/_assets/temp.gdb/DataSources')
    
            if not exists('_assets/established_link.sde'):
                try:
                    print("Establishing Connection...")
                    management.CreateDatabaseConnection('_assets','established_link',init_vars.sde_info[0],instance=init_vars.sde_info[1],account_authentication=init_vars.sde_info[2],database=init_vars.sde_info[3])
                except Exception:
                    isConnected=False
            else:
                try:
                    env.workspace = '_assets/temp.gdb'
                    if Exists('DataSources'):
                        mgnt_Delete('DataSources')
                    try:
                        env.workspace = '_assets/established.sde'
                    except Exception:
                        isConnected = False
                        remove('_assets/established_link.sde')
                        try:
                            management.CreateDatabaseConnection('_assets','established_link',init_vars.sde_info[0],instance=init_vars.sde_info[1],account_authentication=init_vars.sde_info[2],database=init_vars.sde_info[3])
                            isConnected = True
                        except Exception:
                            pass
                except Exception:
                    pass
            if isConnected:
                try:
                    print("Shifting workspace to SDE...")
                    env.workspace = '_assets/established_link.sde'
                    print("Copying DGMRgeo.DBO.DataSources to temp.gdb...")
                    management.Copy('DGMRgeo.DBO.DataSources',f"{getcwd()}/_assets/temp.gdb/DataSources")
                    print('Shifting workspace to temp.gdb...')
                    env.workspace = '_assets/temp.gdb'
                    print("Generating Excel file from DataSource table in temp.gdb...")
                    conversion.TableToExcel(f'{getcwd()}/_assets/temp.gdb/DataSources',f"{getcwd()}/DataSources.xlsx")
                    print("Generation successful.")
                    rmtree('_temp')
                except Exception:
                    isConnected = False
                    if exists('DataSource.xlsx'):
                        remove('DataSources.xlsx')
                    copyfile('_temp/DataSources.xlsx','DataSources.xlsx')
                    rmtree('_temp')
    
            if not isConnected:
                if exists('_temp'):
                    if exists('DataSources.xlsx'):
                        copyfile('_temp/DataSources.xlsx','DataSources.xlsx')
                        rmtree('_temp')
    
            if Exists(f'{getcwd()}/_assets/temp.gdb/DataSources'):
                mgnt_Delete(f'{getcwd()}/_assets/temp.gdb/DataSources')

        env.workspace = check_dir[:]

        try:
            if not Exists("DataSources") or not dc_object.datasources_created:
                if not Exists("DataSources"):
                    print("DataSources table missing/incompleted from geodatabase!\nGenerating DataSources table...")
                    mgnt_CreateTable(out_path=env.workspace,out_name='DataSources',out_alias='DataSources')
                    dc_object.datasources_generated = True
                if not 'Source' in (fields := {field.name for field in ListFields('DataSources')}):
                    mgnt_AddField(in_table='DataSources',field_name='Source',field_type='TEXT',field_length=500,field_alias='Source',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                if not 'Notes' in fields:
                    mgnt_AddField(in_table='DataSources',field_name='Notes',field_type='TEXT',field_length=300,field_alias='Notes',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                if not 'URL' in fields:
                    mgnt_AddField(in_table='DataSources',field_name='URL',field_type='TEXT',field_length=300,field_alias='URL',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                if not 'DataSources_ID' in fields:
                    mgnt_AddField(in_table='DataSources',field_name='DataSources_ID',field_type='TEXT',field_length=300,field_alias='DataSources_ID',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                if not 'created_user' in fields:
                    mgnt_AddField(in_table='DataSources',field_name='created_user',field_type='TEXT',field_length=255,field_alias='created_user',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                if not 'created_date' in fields:
                    mgnt_AddField(in_table='DataSources',field_name='created_date',field_type='DATE',field_alias='created_date',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                if not 'last_edited_user' in fields:
                    mgnt_AddField(in_table='DataSources',field_name='last_edited_user',field_type='TEXT',field_length=255,field_alias='last_edited_user',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                if not 'last_edited_date' in fields:
                    mgnt_AddField(in_table='DataSources',field_name='last_edited_date',field_type='DATE',field_alias='last_edited_date',field_is_nullable='NULLABLE',field_is_required='NON_REQUIRED')
                try:
                    mgnt_EET('DataSources','created_user','created_date','last_edited_user','last_edited_date','NO_ADD_FIELDS','UTC')
                except Exception:
                    pass
                del fields
                gc.collect()
                dc_object.datasources_created = True
                if dc_object.explicit_rerun:
                    try:
                        edit.stopOperation()
                    except Exception:
                        pass
                    try:
                        edit.stopEditing(save_changes=True)
                    except Exception:
                        pass
                    edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
                print("Data successfully generated!")
            else:
                dc_object.datasources_created = True

            test_dasids = set()

            def getDASID(item_path : str) -> None:
                if (fields := tuple([field.name for field in ListFields(item_path) if field.name in relevant_fields])):
                    num_fields = len(fields)
                    for row in da.SearchCursor(item_path,fields):
                        for n in range(num_fields):
                            if row[n] is None:
                                continue
                            current_dasid = sub('[^A-Za-z0-9|]+','',row[n])
                            current_dasid = current_dasid.upper()
                            if current_dasid.find('|') != -1:
                                while current_dasid.find('|') != -1:
                                    test_dasids.add(current_dasid[:current_dasid.find('|')])
                                    current_dasid = current_dasid[current_dasid.find('|')+1:]
                                test_dasids.add(current_dasid)
                            else:
                                test_dasids.add(current_dasid)
                return None

            print("Obtaining DASID values from geodatabase from all tables and feature classes excluding DataSources table...")

            for a in range(len(gdb_info.datasets)):
                for b in range(len(gdb_info.pnt_fc_names[a])):
                    getDASID(f'{gdb_info.datasets[a]}/{gdb_info.pnt_fc_names[a][b]}')
                for b in range(len(gdb_info.pln_fc_names[a])):
                    getDASID(f'{gdb_info.datasets[a]}/{gdb_info.pln_fc_names[a][b]}')
                for b in range(len(gdb_info.poly_fc_names[a])):
                    getDASID(f'{gdb_info.datasets[a]}/{gdb_info.poly_fc_names[a][b]}')

            if 'DataSources' in (temp_tables := list(gdb_info.tables)):
                temp_tables.remove('DataSources')
            if 'GeoMaterialDict' in temp_tables:
                temp_tables.remove('GeoMaterialDict')

            for table in temp_tables:
                getDASID(table)

            del temp_tables
            gc.collect()

            print("Data successfully obtained!\nGetting relevant data from master DataSources table and writing to DataSources table in geodatabase...")

            wb = load_workbook('DataSources.xlsx')
            sh = wb[wb.sheetnames[0]]
            valid_dasids = dict()
            valid_keys = []
            for n in range(sh.max_row-1):
                if (dasid := sh['E%i' % (n+2)].value) in test_dasids:
                    valid_dasids[dasid] = (sh['B%i' % (n+2)].value,sh['C%i' % (n+2)].value,sh['D%i' % (n+2)].value)
                    valid_keys.append(dasid)
            valid_keys = tuple({f'DAS{item}' for item in sorted([valid_key[3:] for valid_key in valid_keys])})
            if dc_object.datasources_generated:
                with da.InsertCursor('DataSources',['Source','Notes','URL','DataSources_ID']) as cursor:
                    for n in range(len(valid_keys)):
                        cursor.insertRow((valid_dasids[valid_keys[n]][0],valid_dasids[valid_keys[n]][1],valid_dasids[valid_keys[n]][2],valid_keys[n]))
            else:
                current_rows = 0
                for row in da.SearchCursor('DataSources',['DataSources_ID']):
                    current_rows += 1
                if (new_rows := len(valid_keys) - current_rows) > 0:
                    with da.InsertCursor('DataSources',['Source','Notes','URL','DataSources_ID']) as cursor:
                        for n in range(new_rows):
                            cursor.insertRow((None,None,None,None))
                max_counter = len(valid_keys)
                counter = 0
                with da.UpdateCursor('DataSources',['Source','Notes','URL','DataSources_ID']) as cursor:
                    for row in cursor:
                        if counter == max_counter:
                            break
                        for n in range(3):
                            row[n] = valid_dasids[valid_keys[counter]][n]
                        row[3] = valid_keys[counter]
                        cursor.updateRow(row)
                        counter += 1

            print("Data successfully written to DataSources table in geodatabase!")

            if dc_object.explicit_rerun:
                try:
                    edit.stopOperation()
                except Exception:
                    pass
                try:
                    edit.stopEditing(save_changes=True)
                except Exception:
                    pass
        except Exception:
            if not dc_object.explicit_rerun:
                print("Denied editing permissions for DataSources table.\nRerunning with explicit editing requests...\n")
                dc_object.explicit_rerun = True ; gc.collect() ; autopopulateDAS()
            else:
                editorErrorMessage()

        dc_object.explicit_rerun = False

        return

    def selection_boxes(entry_string : str):
        # W A R N I N G ! W A R N I N G ! W A R N I N G !
        # This will result in a very minor memory leak due to how tkinter works.
        # A means of preventing this has yet to be discovered but is expected to
        # exist.

        root.withdraw()

        num_datasets = len(gdb_info.datasets)
        dc_object.box_vals = [None for n in range(16)]

        root2 = Tk()
        root2.geometry('300x200')
        root2.resizable(False,False)
        root2.title('Selection')

        box_var = [StringVar(root2) for n in range(16)]

        def updateBoxes(num):
            dc_object.append_box_vals(box_var[num].get(),num)

        def destroy_sub():
            gen_master(entry_string)
            root.deiconify()
            root2.destroy()

        title_label = Label(root2,text="Select Datasets to add FaultDecorations").pack()

        ttk_Checkbox(root2,text=gdb_info.datasets[0],command=partial(updateBoxes,0),variable=box_var[0],onvalue=gdb_info.datasets[0],offvalue='NIHIL').pack()

        if num_datasets >= 2:
            ttk_Checkbox(root2,text=gdb_info.datasets[1],command=partial(updateBoxes,1),variable=box_var[1],onvalue=gdb_info.datasets[1],offvalue='NIHIL').pack()
        if num_datasets >= 3:
            ttk_Checkbox(root2,text=gdb_info.datasets[2],command=partial(updateBoxes,2),variable=box_var[2],onvalue=gdb_info.datasets[2],offvalue='NIHIL').pack()
        if num_datasets >= 4:
            ttk_Checkbox(root2,text=gdb_info.datasets[3],command=partial(updateBoxes,3),variable=box_var[3],onvalue=gdb_info.datasets[3],offvalue='NIHIL').pack()
        if num_datasets >= 5:
            ttk_Checkbox(root2,text=gdb_info.datasets[4],command=partial(updateBoxes,4),variable=box_var[4],onvalue=gdb_info.datasets[4],offvalue='NIHIL').pack()
        if num_datasets >= 6:
            ttk_Checkbox(root2,text=gdb_info.datasets[5],command=partial(updateBoxes,5),variable=box_var[5],onvalue=gdb_info.datasets[5],offvalue='NIHIL').pack()
        if num_datasets >= 7:
            ttk_Checkbox(root2,text=gdb_info.datasets[6],command=partial(updateBoxes,6),variable=box_var[6],onvalue=gdb_info.datasets[6],offvalue='NIHIL').pack()
        if num_datasets >= 8:
            ttk_Checkbox(root2,text=gdb_info.datasets[7],command=partial(updateBoxes,7),variable=box_var[7],onvalue=gdb_info.datasets[7],offvalue='NIHIL').pack()
        if num_datasets >= 9:
            ttk_Checkbox(root2,text=gdb_info.datasets[8],command=partial(updateBoxes,8),variable=box_var[8],onvalue=gdb_info.datasets[8],offvalue='NIHIL').pack()
        if num_datasets >= 10:
            ttk_Checkbox(root2,text=gdb_info.datasets[9],command=partial(updateBoxes,9),variable=box_var[9],onvalue=gdb_info.datasets[9],offvalue='NIHIL').pack()
        if num_datasets >= 11:
            ttk_Checkbox(root2,text=gdb_info.datasets[10],command=partial(updateBoxes,10),variable=box_var[10],onvalue=gdb_info.datasets[10],offvalue='NIHIL').pack()
        if num_datasets >= 12:
            ttk_Checkbox(root2,text=gdb_info.datasets[11],command=partial(updateBoxes,11),variable=box_var[11],onvalue=gdb_info.datasets[11],offvalue='NIHIL').pack()
        if num_datasets >= 13:
            ttk_Checkbox(root2,text=gdb_info.datasets[12],command=partial(updateBoxes,12),variable=box_var[12],onvalue=gdb_info.datasets[12],offvalue='NIHIL').pack()
        if num_datasets >= 14:
            ttk_Checkbox(root2,text=gdb_info.datasets[13],command=partial(updateBoxes,13),variable=box_var[13],onvalue=gdb_info.datasets[13],offvalue='NIHIL').pack()
        if num_datasets >= 15:
            ttk_Checkbox(root2,text=gdb_info.datasets[14],command=partial(updateBoxes,14),variable=box_var[14],onvalue=gdb_info.datasets[14],offvalue='NIHIL').pack()
        if num_datasets >= 16:
            ttk_Checkbox(root2,text=gdb_info.datasets[15],command=partial(updateBoxes,15),variable=box_var[15],onvalue=gdb_info.datasets[15],offvalue='NIHIL').pack()
        Button(root2,text='Confirm',command=destroy_sub).pack()

        root2.mainloop()

        return

    def gen_master(entry_string : str):

        env.workspace = check_dir[:]

        def terminal_string():
            if 16 - dc_object.box_vals.count(None) == 1:
                print("Generating %s..." % entry_string)
            else:
                print("Generating %i instances of %s..." % (16-dc_object.box_vals.count(None),entry_string))

        if dc_object.explicit_rerun:
            edit = da.Editor(env.workspace) ; edit.startEditing(with_undo=False,multiuser_mode=False) ; edit.startOperation()
        else:
            terminal_string()
            try:
                if entry_string == 'FaultDecorations':
                    for a in range(len((relevant_datasets := tuple([item for item in dc_object.box_vals if not item is None])))):
                        prefix = ''
                        if relevant_datasets[a].startswith("CrossSection"):
                            prefix = f"CS{relevant_datasets[a][-1]}"
                        if (now_f_nomin := f'{prefix}FaultDecorations') in gdb_info.pnt_fc_names[a]:
                            print("\n%s already exists in %s!\n" % (now_f_nomin,relevant_datasets[a]))
                            continue
                        mgnt_CreateFC(out_path=relevant_datasets[a],out_name=now_f_nomin,geometry_type="POINT",spatial_reference=gdb_info.spatial_reference[a])
                        current_fc = f'{relevant_datasets[a]}/{now_f_nomin}'
                        if not 'Symbol' in (fields := {field.name for field in ListFields(current_fc)}):
                            mgnt_AddField(in_table=current_fc,field_name='Symbol',field_type='TEXT',field_length=254,field_alias='Symbol',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'PlotAtScale' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='PlotAtScale',field_type='FLOAT',field_alias='PlotAtScale',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'DataSourceID' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='DataSourceID',field_type='TEXT',field_length=50,field_alias='DataSourceID',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'Notes' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='Notes',field_type='TEXT',field_length=254,field_alias='Notes',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'created_user' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='created_user',field_type='TEXT',field_length=255,field_alias='created_user',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'created_date' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='created_date',field_type='DATE',field_alias='created_date',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'last_edited_user' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='last_edited_user',field_type='TEXT',field_length=255,field_alias='last_edited_user',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'last_edited_date' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='last_edited_date',field_type='DATE',field_alias='last_edited_date',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not 'Azimuth' in fields:
                            mgnt_AddField(in_table=current_fc,field_name='Azimuth',field_type='FLOAT',field_alias='Azimuth',field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        if not '%s_ID' % now_f_nomin in fields:
                            mgnt_AddField(in_table=current_fc,field_name='%s_ID' % now_f_nomin,field_type='TEXT',field_length=50,field_alias='%s_ID' % now_f_nomin,field_is_nullable="NULLABLE",field_is_required="NON_REQUIRED")
                        mgnt_EET(current_fc,'created_user','created_date','last_edited_user','last_edited_date','NO_ADD_FIELDS','UTC')
                    if dc_object.explicit_rerun:
                        try:
                            edit.stopOperation()
                        except Exception:
                            pass
                        try:
                            edit.stopEditing(save_changes=True)
                        except Exception:
                            pass
            except Exception:
                if not dc_object.explicit_rerun:
                    print("Permissions to generate %s were denied.\nRerunning process with explicit editing permissions...\n" % entry_string)
                    dc_object.explicit_rerun = True ; gc.collect() ; gen_master(entry_string)
                else:
                    editorErrorMessage()

        print("Process successful!\n")

        dc_object.explicit_rerun = False

        gc.collect()

    def gen_backup():

        loop_variable = True
        zip_dir_name = f'{dir_name}.zip'
        base_dir = check_dir.rstrip(dir_name)
        base_dir = base_dir.rstrip('/')

        temp_root = Tk() ; temp_root.withdraw()

        while loop_variable:
            user_response = None
            if not exists((backup_loc := filedialog.askdirectory(title='Select Backup Location'))):
                error_message('Unknown Directory/Folder',"Specified directory/folder cannot be found or a directory/folder was not selected.")
            else:
                if zip_dir_name in set(listdir(backup_loc)):
                    if (user_response := yesno_message("%s Already Exists","Would you like to overwrite this existing ZIP file?")) == 'yes':
                        remove(f'{backup_loc}/{zip_dir_name}')
                    else:
                        print("\nPlease choose a different location!\n")
                if user_response in (None,'yes'):
                    current_dir = getcwd()
                    chdir(backup_loc)
                    with ZipFile(zip_dir_name,'w',compression=ZIP_DEFLATED) as zf:
                        chdir(check_dir)
                        for item in tuple(listdir()):
                            zf.write(item)
                    chdir(current_dir)
                    dc_object.backup_gdb = f'{backup_loc}/{zip_dir_name}'
                break

        temp_root.update() ; temp_root.destroy()

        del loop_variable ; del current_dir ; del zip_dir_name ; del base_dir ; del backup_loc ; del user_response
        gc.collect()

    def metadata_get_counties():

        pass

    label_auto = Label(root,text="Auto-Complete Tools",font=helv_label,bg='lightblue').place(relx=0.5,rely=0.055,anchor='n')

    btn_clear_and_reset_id_fields = Button(root,text="Clear and Reset Identifier Fields",font=helv_button,command=resetFcId,bg='lightgrey',activebackground='grey').place(relx=0.5,rely=0.11,anchor='n')

    btn_mapunit = Button(root,text="Fill MapUnit Field for\nPoint Feature Classes",command=point_mapunit,font=helv_button,bg='lightgrey',activebackground='grey').place(relx=0.5,rely=0.155,anchor='n')

    btn_alpha_gloss = Button(root,text="Alphabetize Glossary Table",command=alphabetizeGlossary,font=helv_button,bg='lightgrey',activebackground='grey').place(relx=0.5,rely=0.22,anchor='n')

    btn_auto_das = Button(root,text="Autopopulate DataSources Table",command=autopopulateDAS,font=helv_button,bg='lightgrey',activebackground='grey').place(relx=0.5,rely=0.265,anchor='n')

    btn_remove_ET = Button(root,text="Disable Editor Tracking\n and Delete Related Fields",font=helv_button,command=removeETfields,bg='lightgrey',activebackground='grey').place(relx=0.5,rely=0.31,anchor='n')

    label_new_gen = Label(root,text="New-Gen. Tools",font=helv_label,bg='lightblue').place(relx=0.5,rely=0.455,anchor='n')

    btn_backup_gdb = Button(root,text="Create Geodatabase Backup",command=gen_backup,font=helv_button,bg='lightgrey',activebackground='grey').place(relx=0.5,rely=0.51,anchor='n')

    btn_gen_faultdecorations = Button(root,text="FaultDecorations",font=helv_button,command=partial(selection_boxes,'FaultDecorations'),bg='lightgrey',activebackground='grey').place(relx=0.5,rely=0.555,anchor='n')

    btn_switch_gdb = Button(root,text="Switch Selected\nGeodatabase",font=helv_button,command=switchGDB,bg='royalblue',activebackground='grey').place(relx=0.9,rely=0.9,anchor='n')

    root.mainloop()

    return None


if __name__ == '__main__':

    main()
